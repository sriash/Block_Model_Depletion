"""
exporter.py
Exports the DIS pivot table and audit log to CSV and formatted Excel.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Return DataFrame as UTF-8 encoded CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")


def to_excel_bytes(
    pivot_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    audit_df: pd.DataFrame | None = None,
) -> bytes:
    """
    Build a formatted Excel workbook with:
      - Sheet 1: DIS Pivot Table (formatted)
      - Sheet 2: Raw Data
      - Sheet 3: Audit Log (if provided)
    Returns bytes suitable for st.download_button.
    """
    wb = Workbook()

    _write_pivot_sheet(wb.active, pivot_df)
    wb.active.title = "DIS Pivot Table"

    raw_sheet = wb.create_sheet("Raw Data")
    _write_plain_sheet(raw_sheet, raw_df)

    if audit_df is not None and not audit_df.empty:
        audit_sheet = wb.create_sheet("Audit Log")
        _write_plain_sheet(audit_sheet, audit_df)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_pivot_sheet(ws, df: pd.DataFrame):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill("solid", fgColor="D6E4F0")
    total_font = Font(bold=True, size=11)
    border = Border(
        bottom=Side(style="thin"),
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
    )

    ws.append(["Depletion Information System (DIS) — Pivot Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if r_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            elif r_idx == ws.max_row:
                cell.fill = total_fill
                cell.font = total_font

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)


def _write_plain_sheet(ws, df: pd.DataFrame):
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
